"""日志导出面板 - 将日志导出为 TXT/CSV/HTML 格式"""

import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime


class ExportPanel(ttk.LabelFrame):
    """日志导出面板"""

    def __init__(self, parent, log_panel=None):
        super().__init__(parent, text=' 日志导出 ', padding=8)
        self._log_panel = log_panel
        self._build_ui()

    def _build_ui(self):
        # 导出范围
        row1 = ttk.Frame(self)
        row1.pack(fill=tk.X, pady=(0, 4))
        
        ttk.Label(row1, text='导出范围:').pack(side=tk.LEFT)
        self.range_var = tk.StringVar(value='全部日志')
        ttk.Combobox(row1, textvariable=self.range_var,
                     values=['全部日志', '仅TX', '仅RX', '当前选中'],
                     state='readonly', width=12).pack(side=tk.LEFT, padx=(4, 0))

        # 导出格式
        row2 = ttk.Frame(self)
        row2.pack(fill=tk.X, pady=(0, 4))
        
        ttk.Label(row2, text='导出格式:').pack(side=tk.LEFT)
        self.format_var = tk.StringVar(value='TXT')
        for fmt in ['TXT', 'CSV', 'HTML', 'Excel', 'PDF']:
            ttk.Radiobutton(row2, text=fmt, variable=self.format_var,
                            value=fmt).pack(side=tk.LEFT, padx=(4, 2))

        # 选项
        row3 = ttk.Frame(self)
        row3.pack(fill=tk.X, pady=(0, 4))
        
        self.include_time_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text='包含时间戳', variable=self.include_time_var).pack(side=tk.LEFT, padx=(0, 10))
        
        self.include_raw_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(row3, text='包含原始数据', variable=self.include_raw_var).pack(side=tk.LEFT)

        # 导出按钮
        row4 = ttk.Frame(self)
        row4.pack(fill=tk.X, pady=(0, 4))
        
        ttk.Button(row4, text='📥 导出', command=self._export, width=12).pack(side=tk.LEFT)

        # 状态
        self.status_var = tk.StringVar(value='')
        ttk.Label(self, textvariable=self.status_var, foreground='gray').pack(anchor=tk.W)

    def _export(self):
        """导出日志"""
        if not self._log_panel:
            messagebox.showwarning('提示', '日志面板未连接')
            return
        
        # 选择保存路径
        fmt = self.format_var.get().lower()
        file_path = filedialog.asksaveasfilename(
            title='导出日志',
            defaultextension=f'.{fmt}',
            filetypes=[
                (f'{fmt.upper()} 文件', f'*.{fmt}'),
                ('所有文件', '*.*'),
            ])
        if not file_path:
            return
        
        try:
            # 获取日志内容
            log_text = self._log_panel.text.get('1.0', tk.END)

            if fmt == 'txt':
                self._export_txt(file_path, log_text)
            elif fmt == 'csv':
                self._export_csv(file_path, log_text)
            elif fmt == 'html':
                self._export_html(file_path, log_text)
            elif fmt == 'excel':
                self._export_excel(file_path, log_text)
            elif fmt == 'pdf':
                self._export_pdf(file_path, log_text)

            self.status_var.set(f'✅ 已导出: {os.path.basename(file_path)}')
        except Exception as e:
            messagebox.showerror('导出失败', str(e))
            self.status_var.set(f'❌ 导出失败: {e}')

    def _export_txt(self, file_path: str, content: str):
        """导出为 TXT"""
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write('=' * 60 + '\n')
            f.write(f'通信日志导出 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
            f.write('=' * 60 + '\n\n')
            f.write(content)

    def _export_csv(self, file_path: str, content: str):
        """导出为 CSV"""
        import csv
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['时间', '方向', '数据'])
            
            for line in content.split('\n'):
                if not line.strip():
                    continue
                # 解析日志行格式: [时间] TX/RX >> 数据
                if line.startswith('['):
                    try:
                        time_end = line.index(']')
                        timestamp = line[1:time_end]
                        rest = line[time_end + 1:].strip()
                        
                        if 'TX' in rest:
                            direction = 'TX'
                            data = rest[rest.index('>>') + 2:].strip()
                        elif 'RX' in rest:
                            direction = 'RX'
                            data = rest[rest.index('<<') + 2:].strip()
                        else:
                            direction = 'INFO'
                            data = rest
                        
                        writer.writerow([timestamp, direction, data])
                    except (ValueError, IndexError):
                        writer.writerow(['', '', line.strip()])

    def _export_html(self, file_path: str, content: str):
        """导出为 HTML"""
        html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<title>通信日志导出</title>
<style>
body {{ font-family: 'Courier New', monospace; background: #1e1e1e; color: #d4d4d4; padding: 20px; }}
h1 {{ color: #569cd6; }}
.log {{ white-space: pre-wrap; font-size: 13px; line-height: 1.5; }}
.tx {{ color: #4fc1ff; }}
.rx {{ color: #6a9955; }}
.info {{ color: #ce9178; }}
.time {{ color: #808080; }}
.header {{ color: #569cd6; border-bottom: 1px solid #333; padding-bottom: 10px; }}
</style>
</head>
<body>
<div class="header">
<h1>📡 通信日志导出</h1>
<p>导出时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}</p>
</div>
<div class="log">
'''
        for line in content.split('\n'):
            if not line.strip():
                html += '<br>\n'
                continue
            
            css_class = 'info'
            if 'TX' in line:
                css_class = 'tx'
            elif 'RX' in line:
                css_class = 'rx'
            
            # 高亮时间戳
            import re
            line = re.sub(r'(\[\d{2}:\d{2}:\d{2}\.\d+\])',
                         r'<span class="time">\1</span>', line)
            
            html += f'<span class="{css_class}">{line}</span>\n'
        
        html += '''</div>
</body>
</html>'''
        
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html)

    def _export_excel(self, file_path: str, content: str):
        """导出为 Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = '通信日志'

        # 标题行
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = f'通信日志导出 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        title_cell.font = Font(size=14, bold=True, color='1E90FF')
        title_cell.alignment = Alignment(horizontal='center')

        # 表头
        headers = ['时间', '方向', '数据']
        header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # 数据行
        row = 4
        tx_font = Font(color='4FC1FF')
        rx_font = Font(color='6A9955')
        info_font = Font(color='CE9178')

        for line in content.split('\n'):
            if not line.strip():
                continue
            timestamp = ''
            direction = ''
            data = line.strip()

            if line.startswith('['):
                try:
                    time_end = line.index(']')
                    timestamp = line[1:time_end]
                    rest = line[time_end + 1:].strip()
                    if 'TX' in rest and '>>' in rest:
                        direction = 'TX'
                        data = rest[rest.index('>>') + 2:].strip()
                    elif 'RX' in rest and '<<' in rest:
                        direction = 'RX'
                        data = rest[rest.index('<<') + 2:].strip()
                    else:
                        direction = 'INFO'
                        data = rest
                except (ValueError, IndexError):
                    pass

            ws.cell(row=row, column=1, value=timestamp)
            ws.cell(row=row, column=2, value=direction)
            data_cell = ws.cell(row=row, column=3, value=data)

            if direction == 'TX':
                data_cell.font = tx_font
            elif direction == 'RX':
                data_cell.font = rx_font
            else:
                data_cell.font = info_font

            row += 1

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 80
        wb.save(file_path)

    def _export_pdf(self, file_path: str, content: str):
        """导出为 PDF"""
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, HRFlowable
        from reportlab.lib.colors import HexColor

        doc = SimpleDocTemplate(file_path, pagesize=A4,
                                topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('Title', parent=styles['Heading1'],
                                     fontSize=16, textColor=HexColor('#1E90FF'),
                                     spaceAfter=10)
        normal_style = ParagraphStyle('Log', parent=styles['Code'],
                                      fontSize=8, leading=12,
                                      spaceAfter=2)

        elements = []
        elements.append(Paragraph(
            f'通信日志导出 - {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
            title_style))
        elements.append(HRFlowable(width='100%', thickness=1, color=HexColor('#333333')))
        elements.append(Spacer(1, 10))

        for line in content.split('\n'):
            if not line.strip():
                continue
            escaped = line.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            escaped = escaped.replace('\n', '<br/>')
            elements.append(Paragraph(escaped, normal_style))

        doc.build(elements)

    def get_settings(self) -> dict:
        return {
            'range': self.range_var.get(),
            'format': self.format_var.get(),
            'include_time': self.include_time_var.get(),
            'include_raw': self.include_raw_var.get(),
        }

    def load_settings(self, settings: dict):
        if not settings:
            return
        if 'range' in settings:
            self.range_var.set(settings['range'])
        if 'format' in settings:
            self.format_var.set(settings['format'])
        if 'include_time' in settings:
            self.include_time_var.set(settings['include_time'])
        if 'include_raw' in settings:
            self.include_raw_var.set(settings['include_raw'])
